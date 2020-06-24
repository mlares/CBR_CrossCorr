# load_ext autoreload
# autoreload 2

from openpyxl import load_workbook
from collections import namedtuple
import jinja2
from os.path import join as pathjoin

wb = load_workbook('SYSH_suite.xlsx')
print(wb.sheetnames)

first_column = 'A'
last_column = 'O'
col_range = f"{first_column}:{last_column}"

suite = wb['SAYSHELL']
s=suite[col_range]
 
for row in suite.iter_rows(min_row=1, max_row=1):
    names = []
    nms = ''
    for cell in row:
        names.append(cell.value)
        nms = nms + str(cell.value) + ' '
    print(names)
    nms = nms.strip()

parset = namedtuple('pars', nms)



source_dir = './'
template_file = 'config_template.ini'
templateLoader = jinja2.FileSystemLoader(searchpath=source_dir)

latex_jinja_env = jinja2.Environment(
    block_start_string=r"\BLOCK{",
    block_end_string='}',
    variable_start_string=r'\VAR{',
    variable_end_string='}',
    comment_start_string=r'\#{',
    comment_end_string='}',
    line_statement_prefix='%%',
    line_comment_prefix='%#',
    trim_blocks=True,
    autoescape=False,
    loader=templateLoader
)
template_intro = latex_jinja_env.get_template(template_file)


k = 0
for row in suite.iter_rows(min_row=2, max_row=9):
    r = []
    k += 1
    for cell in row:
        r.append(str(cell.value).strip())

    t = tuple(r)
    p = parset(*t)

    ininame = 'config_SYSH_' + str(k).zfill(3) + '.ini'

    filename = pathjoin(source_dir, ininame)
    target = open(filename, 'w')
    target.write(template_intro.render(p=p))
    target.close()





